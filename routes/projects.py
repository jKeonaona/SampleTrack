import io
import json
import os
import tempfile
from datetime import date, datetime

from flask import Blueprint, Response, render_template, request, redirect, url_for, flash
from flask_login import login_required
from sqlalchemy import or_

from models import (
    AirMonitorReport,
    FieldSampleRecord,
    Project,
    Result,
    Sample,
    SamplingEvent,
    Threshold,
    db,
)
from parsers.lab_report import MATRIX_OPTIONS, parse_lab_report
from routes._helpers import (
    FLAG_DISPLAY,
    csv_response,
    safe_filename_part,
    _threshold_labels,
)
from utils.calculations import evaluate_result, project_status_summary, worst_sample_status

projects_bp = Blueprint("projects", __name__, url_prefix="/projects")


@projects_bp.route("/", methods=["GET"])
@login_required
def list_projects():
    show_archived = request.args.get("show_archived") == "1"
    query = Project.query
    if not show_archived:
        query = query.filter(Project.status != "archived")
    projects = query.order_by(Project.created_at.desc()).all()
    return render_template(
        "projects/list.html",
        projects=projects,
        show_archived=show_archived,
    )


@projects_bp.route("/new", methods=["GET"])
@login_required
def new():
    return render_template("projects/new.html", form={})


@projects_bp.route("/", methods=["POST"])
@login_required
def create():
    form = {
        "project_number": (request.form.get("project_number") or "").strip(),
        "name": (request.form.get("name") or "").strip(),
        "client": (request.form.get("client") or "").strip(),
        "location": (request.form.get("location") or "").strip(),
        "status": (request.form.get("status") or "active").strip(),
        "jurisdiction": (request.form.get("jurisdiction") or "California").strip(),
    }

    if not form["project_number"]:
        flash("Project number is required.", "error")
        return render_template("projects/new.html", form=form), 400
    if not form["name"]:
        flash("Name is required.", "error")
        return render_template("projects/new.html", form=form), 400
    if form["jurisdiction"] not in VALID_JURISDICTIONS:
        flash(f"Jurisdiction must be one of: {', '.join(VALID_JURISDICTIONS)}.", "error")
        return render_template("projects/new.html", form=form), 400

    existing = Project.query.filter_by(project_number=form["project_number"]).first()
    if existing is not None:
        flash(f"Project number '{form['project_number']}' already exists.", "error")
        return render_template("projects/new.html", form=form), 400

    project = Project(
        project_number=form["project_number"],
        name=form["name"],
        client=form["client"] or None,
        location=form["location"] or None,
        status=form["status"] or "active",
        jurisdiction=form["jurisdiction"],
    )
    db.session.add(project)
    db.session.commit()

    flash(f"Project '{project.project_number}' created.", "success")
    return redirect(url_for("projects.detail", project_id=project.id))


@projects_bp.route("/<int:project_id>", methods=["GET"])
@login_required
def detail(project_id):
    project = Project.query.get_or_404(project_id)
    samples = (
        Sample.query
        .filter_by(project_id=project.id)
        .order_by(Sample.collection_date.desc(), Sample.client_sample_id.asc())
        .all()
    )
    sample_statuses = {s.id: worst_sample_status(s) for s in samples}
    project_summary = project_status_summary(samples)
    return render_template(
        "projects/detail.html",
        project=project,
        samples=samples,
        sample_statuses=sample_statuses,
        project_summary=project_summary,
    )


@projects_bp.route("/<int:project_id>/export", methods=["GET"])
@login_required
def export_project(project_id):
    project = Project.query.get_or_404(project_id)
    samples = (
        Sample.query
        .filter_by(project_id=project.id)
        .order_by(Sample.collection_date.desc(), Sample.client_sample_id.asc())
        .all()
    )

    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    safe_number = safe_filename_part(project.project_number)
    filename = f"project_{safe_number}_export_{timestamp}.csv"
    return csv_response(samples, filename)


@projects_bp.route("/<int:project_id>/archive", methods=["POST"])
@login_required
def archive(project_id):
    project = Project.query.get_or_404(project_id)
    project.status = "archived"
    project.archived_at = datetime.utcnow()
    db.session.commit()
    flash("Project archived.", "success")
    return redirect(url_for("projects.detail", project_id=project.id))


@projects_bp.route("/<int:project_id>/unarchive", methods=["POST"])
@login_required
def unarchive(project_id):
    project = Project.query.get_or_404(project_id)
    project.status = "active"
    project.archived_at = None
    db.session.commit()
    flash("Project unarchived.", "success")
    return redirect(url_for("projects.detail", project_id=project.id))


@projects_bp.route("/<int:project_id>/export/excel", methods=["GET"])
@login_required
def export_project_excel(project_id):
    project = Project.query.get_or_404(project_id)
    return _build_project_excel_response(project)


VALID_PROJECT_STATUSES = ("active", "archived", "complete")
VALID_JURISDICTIONS = ("California", "Federal")


@projects_bp.route("/<int:project_id>/edit", methods=["GET"])
@login_required
def edit(project_id):
    project = Project.query.get_or_404(project_id)
    form = {
        "project_number": project.project_number,
        "name": project.name or "",
        "client": project.client or "",
        "location": project.location or "",
        "status": project.status or "active",
        "jurisdiction": project.jurisdiction or "California",
    }
    return render_template("projects/edit.html", project=project, form=form)


@projects_bp.route("/<int:project_id>/edit", methods=["POST"])
@login_required
def edit_save(project_id):
    project = Project.query.get_or_404(project_id)
    form = {
        "project_number": (request.form.get("project_number") or "").strip(),
        "name": (request.form.get("name") or "").strip(),
        "client": (request.form.get("client") or "").strip(),
        "location": (request.form.get("location") or "").strip(),
        "status": (request.form.get("status") or "active").strip(),
        "jurisdiction": (request.form.get("jurisdiction") or "California").strip(),
    }

    if not form["project_number"]:
        flash("Project number is required.", "error")
        return render_template("projects/edit.html", project=project, form=form), 400

    if form["status"] not in VALID_PROJECT_STATUSES:
        flash(f"Status must be one of: {', '.join(VALID_PROJECT_STATUSES)}.", "error")
        return render_template("projects/edit.html", project=project, form=form), 400

    if form["jurisdiction"] not in VALID_JURISDICTIONS:
        flash(f"Jurisdiction must be one of: {', '.join(VALID_JURISDICTIONS)}.", "error")
        return render_template("projects/edit.html", project=project, form=form), 400

    if form["project_number"] != project.project_number:
        clash = Project.query.filter_by(project_number=form["project_number"]).first()
        if clash is not None and clash.id != project.id:
            flash(f"Project number '{form['project_number']}' is already in use.", "error")
            return render_template("projects/edit.html", project=project, form=form), 400

    project.project_number = form["project_number"]
    project.name = form["name"] or "Unnamed Project"
    project.client = form["client"] or None
    project.location = form["location"] or None
    project.status = form["status"]
    project.jurisdiction = form["jurisdiction"]
    db.session.commit()

    flash("Project updated.", "success")
    return redirect(url_for("projects.detail", project_id=project.id))


@projects_bp.route("/<int:project_id>/upload", methods=["GET"])
@login_required
def upload_new(project_id):
    project = Project.query.get_or_404(project_id)
    return render_template("uploads/new.html", project=project)


@projects_bp.route("/<int:project_id>/upload", methods=["POST"])
@login_required
def upload(project_id):
    project = Project.query.get_or_404(project_id)

    uploaded = request.files.get("report_pdf")
    if uploaded is None or not uploaded.filename:
        flash("Please choose a PDF to upload.", "error")
        return redirect(url_for("projects.upload_new", project_id=project.id))
    if not uploaded.filename.lower().endswith(".pdf"):
        flash("Only PDF files are accepted.", "error")
        return redirect(url_for("projects.upload_new", project_id=project.id))

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
    os.close(tmp_fd)
    try:
        uploaded.save(tmp_path)
        parser_data = parse_lab_report(tmp_path)
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    parser_data_json = json.dumps(parser_data)
    return render_template(
        "uploads/confirm.html",
        project=project,
        parser_data=parser_data,
        parser_data_json=parser_data_json,
        matrix_options=MATRIX_OPTIONS,
    )


@projects_bp.route("/<int:project_id>/upload/save", methods=["POST"])
@login_required
def upload_save(project_id):
    project = Project.query.get_or_404(project_id)

    raw = request.form.get("parser_data") or ""
    try:
        parser_data = json.loads(raw)
    except json.JSONDecodeError:
        flash("Could not read parsed data. Please re-upload the PDF.", "error")
        return redirect(url_for("projects.upload_new", project_id=project.id))

    workorder = parser_data.get("workorder")
    samples_in = parser_data.get("samples") or []

    sample_count = 0
    result_count = 0
    for idx, sample_data in enumerate(samples_in):
        matrix = (request.form.get(f"matrix_{idx}") or sample_data.get("matrix") or "Other").strip()

        sample = Sample(
            project_id=project.id,
            client_sample_id=(sample_data.get("client_sample_id") or sample_data.get("lab_sample_id") or "UNKNOWN"),
            lab_sample_id=sample_data.get("lab_sample_id"),
            lab_workorder=workorder,
            matrix=matrix,
            matrix_code=sample_data.get("matrix_code"),
            is_blank=bool(sample_data.get("is_blank")),
            collection_date=_parse_date(sample_data.get("collection_date")),
            collection_time=sample_data.get("collection_time"),
            sample_volume=sample_data.get("sample_volume"),
        )
        db.session.add(sample)
        db.session.flush()
        sample_count += 1

        for r_data in sample_data.get("results") or []:
            analyte = (r_data.get("analyte") or "").strip()
            result_value = (r_data.get("result_value") or "").strip()
            if not analyte or not result_value:
                continue
            db.session.add(Result(
                sample_id=sample.id,
                analyte=analyte,
                result_value=result_value,
                result_numeric=_parse_numeric(result_value),
                result_units=r_data.get("result_units") or sample_data.get("units"),
                reporting_limit=r_data.get("reporting_limit"),
                dilution_factor=r_data.get("dilution_factor"),
                method_reference=sample_data.get("method"),
                lab_report_number=workorder,
                date_analyzed=_parse_datetime(r_data.get("date_analyzed")),
            ))
            result_count += 1

    db.session.commit()
    flash(f"Saved {sample_count} samples and {result_count} results.", "success")
    return redirect(url_for("projects.detail", project_id=project.id))


def _parse_date(value):
    if not value:
        return None
    value = value.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value):
    if not value:
        return None
    value = value.strip()
    for fmt in (
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _parse_numeric(value):
    if value is None:
        return None
    cleaned = str(value).strip().replace(",", "")
    if cleaned == "" or cleaned.upper() == "ND":
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _fmt_cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return v


def _write_sheet(ws, headers, rows, auto_filter=True):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "A2"

    for r in rows:
        ws.append([_fmt_cell(v) for v in r])

    col_widths = [len(h) for h in headers]
    for r in rows:
        for i, v in enumerate(r):
            text = str(_fmt_cell(v))
            if len(text) > col_widths[i]:
                col_widths[i] = len(text)
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(max(w + 2, 10), 60)

    if auto_filter and rows and headers:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


def _build_project_excel_response(project):
    from openpyxl import Workbook

    samples = (
        Sample.query
        .filter_by(project_id=project.id)
        .order_by(Sample.collection_date.asc().nullsfirst(), Sample.client_sample_id.asc())
        .all()
    )
    sample_ids = [s.id for s in samples]
    client_ids = [s.client_sample_id for s in samples if s.client_sample_id]

    events = (
        SamplingEvent.query
        .filter_by(project_id=project.id)
        .order_by(SamplingEvent.event_date.asc().nullsfirst(), SamplingEvent.created_at.asc())
        .all()
    )

    amrs = []
    fsrs = []
    if client_ids:
        amrs = (
            AirMonitorReport.query
            .filter(AirMonitorReport.client_sample_id.in_(client_ids))
            .all()
        )
        fsrs = (
            FieldSampleRecord.query
            .filter(FieldSampleRecord.client_sample_id.in_(client_ids))
            .all()
        )

    results = []
    if sample_ids:
        results = (
            Result.query
            .filter(Result.sample_id.in_(sample_ids))
            .all()
        )

    sample_by_id = {s.id: s for s in samples}
    event_sample_counts = {e.id: 0 for e in events}
    for s in samples:
        if s.sampling_event_id in event_sample_counts:
            event_sample_counts[s.sampling_event_id] += 1

    wb = Workbook()

    # 1. Project Info
    ws = wb.active
    ws.title = "Project Info"
    from openpyxl.styles import Font
    bold = Font(bold=True)
    info_rows = [
        ("Project Number", project.project_number),
        ("Project Name", project.name),
        ("Client", project.client),
        ("Location", project.location),
        ("Status", project.status),
        ("Jurisdiction", project.jurisdiction),
    ]
    if project.status == "archived" and project.archived_at:
        info_rows.append(("Archived At", project.archived_at))
    info_rows.extend([
        ("Created At", project.created_at),
        ("Updated At", project.updated_at),
    ])
    for k, v in info_rows:
        ws.append([k, _fmt_cell(v)])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    # 2. Sampling Events
    ws = wb.create_sheet("Sampling Events")
    _write_sheet(
        ws,
        ["Event ID", "Event Date", "Matrix Code", "Expected Count", "Actual Count",
         "Notes", "Created By", "Created At"],
        [
            [
                e.id, e.event_date, e.matrix_code, e.expected_sample_count,
                event_sample_counts.get(e.id, 0),
                e.notes,
                (e.created_by.name or e.created_by.email) if e.created_by else None,
                e.created_at,
            ]
            for e in events
        ],
    )

    # 3. Samples
    ws = wb.create_sheet("Samples")
    _write_sheet(
        ws,
        ["Sample ID", "Client Sample ID", "Sampling Event ID", "Sequence Number",
         "Matrix", "Matrix Code", "Is Blank", "Collection Date",
         "Collection Start Time", "Collection End Time", "Sample Volume",
         "Pump Flow Rate", "Lab Workorder", "Lab Sample ID", "Employee Name",
         "Work Area", "Task Description", "Created At"],
        [
            [
                s.id, s.client_sample_id, s.sampling_event_id, s.sequence_number,
                s.matrix, s.matrix_code, bool(s.is_blank), s.collection_date,
                s.collection_start_time, s.collection_end_time, s.sample_volume,
                s.pump_flow_rate, s.lab_workorder, s.lab_sample_id, s.employee_name,
                s.work_area, s.task_description, s.created_at,
            ]
            for s in samples
        ],
    )

    # 4. Field Air Monitor Reports
    ws = wb.create_sheet("Field Air Monitor Reports")
    _write_sheet(
        ws,
        ["ID", "Client Sample ID", "Monitoring Date", "Monitoring Phase",
         "Pump Serial", "Cassette Number", "Time Started", "AM/PM Started",
         "Time Stopped", "AM/PM Stopped", "Total Hours", "Total Minutes",
         "Liters/Minute", "Calibrated Before", "Calibrated Before Rate",
         "Calibrated Before By", "Calibrated After", "Calibrated After Rate",
         "Calibrated After By", "Last Calibration Date", "Measured Rate",
         "Used Since Calibration", "Is Personal Sample", "Worker Monitored",
         "Job Duties", "PPE Worn", "Cassette Worn Properly",
         "Dragged Through Abrasive", "Unusual Happened", "Unusual Details",
         "Is Area Sample", "Monitor Location", "Blasting Location",
         "Area Unusual Events", "Weather Conditions", "Temperature Value",
         "Temperature Unit", "Wind Speed/Direction", "Placed By", "Removed By",
         "Laboratory Sent To", "Date Sent to Lab", "Created At"],
        [
            [
                a.id, a.client_sample_id, a.monitoring_date, a.monitoring_phase,
                a.pump_serial, a.cassette_number, a.time_started, a.am_pm_started,
                a.time_stopped, a.am_pm_stopped, a.total_hours, a.total_minutes,
                a.liters_per_minute, bool(a.calibrated_before), a.calibrated_before_rate,
                a.calibrated_before_by, bool(a.calibrated_after), a.calibrated_after_rate,
                a.calibrated_after_by, a.last_calibration_date, a.measured_rate,
                bool(a.used_since_calibration), bool(a.is_personal_sample),
                a.worker_monitored, a.job_duties, a.ppe_worn,
                bool(a.cassette_worn_properly), bool(a.dragged_through_abrasive),
                bool(a.unusual_happened), a.unusual_details, bool(a.is_area_sample),
                a.monitor_location, a.blasting_location, a.area_unusual_events,
                a.weather_conditions, a.temperature_value, a.temperature_unit,
                a.wind_speed_direction, a.placed_by, a.removed_by,
                a.laboratory_sent_to, a.date_sent_to_lab, a.created_at,
            ]
            for a in amrs
        ],
    )

    # 5. Field Sample Records
    ws = wb.create_sheet("Field Sample Records")
    _write_sheet(
        ws,
        ["ID", "Client Sample ID", "Collection Date", "Collection Time",
         "Collected By", "Location Description", "Matrix-Specific Notes",
         "Analytical Methods Requested", "Laboratory Sent To", "Date Sent to Lab",
         "General Notes", "Created At"],
        [
            [
                f.id, f.client_sample_id, f.collection_date, f.collection_time,
                f.collected_by, f.location_description, f.matrix_specific_notes,
                f.analytical_methods_requested, f.laboratory_sent_to, f.date_sent_to_lab,
                f.general_notes, f.created_at,
            ]
            for f in fsrs
        ],
    )

    # 6. Analytical Results
    ws = wb.create_sheet("Analytical Results")
    result_rows = []
    for r in results:
        s = sample_by_id.get(r.sample_id)
        if s is None:
            continue
        ev = evaluate_result(s, r)
        exceeded = _threshold_labels(ev["evaluations"], "exceeded")
        approaching = _threshold_labels(ev["evaluations"], "approaching")
        result_rows.append([
            r.id, r.sample_id, s.client_sample_id, r.analyte, r.result_value,
            r.result_numeric, r.result_units, r.reporting_limit, r.dilution_factor,
            r.method_reference, r.date_analyzed,
            ev.get("comparison_basis") or "", ev.get("comparison_value"),
            FLAG_DISPLAY.get(ev.get("overall_status"), ""),
            "; ".join(exceeded), "; ".join(approaching),
        ])
    _write_sheet(
        ws,
        ["Result ID", "Sample ID", "Client Sample ID", "Analyte", "Result Value",
         "Result Numeric", "Units", "RL", "DF", "Method", "Date Analyzed",
         "Comparison Basis", "Comparison Value", "Flag Status",
         "Thresholds Exceeded", "Thresholds Approaching"],
        result_rows,
    )

    # 7. Thresholds As-Evaluated
    ws = wb.create_sheet("Thresholds As-Evaluated")
    threshold_rows = _thresholds_as_evaluated(project, samples, results)
    _write_sheet(
        ws,
        ["Analyte", "Matrix", "Threshold Name", "Regulatory Body", "Value",
         "Units", "Threshold Type", "Jurisdiction", "Effective Date",
         "Superseded Date", "Source Citation"],
        threshold_rows,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.utcnow().strftime("%Y%m%d")
    safe_number = safe_filename_part(project.project_number)
    safe_name = safe_filename_part(project.name or "project")
    filename = f"{safe_number}_{safe_name}_export_{today}.xlsx"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _thresholds_as_evaluated(project, samples, results):
    """Threshold rows that could have applied to this project's samples.

    For each unique (analyte, matrix) pair from results, find thresholds whose
    effective window overlaps the project's collection-date range and whose
    jurisdiction matches the project (or is 'Both').
    """
    pairs = set()
    sample_by_id = {s.id: s for s in samples}
    for r in results:
        s = sample_by_id.get(r.sample_id)
        if s is None or not r.analyte or not s.matrix:
            continue
        pairs.add((r.analyte, s.matrix))

    if not pairs:
        return []

    collection_dates = [s.collection_date for s in samples if s.collection_date]
    if collection_dates:
        min_iso = min(collection_dates).isoformat()
        max_iso = max(collection_dates).isoformat()
    else:
        today = date.today().isoformat()
        min_iso = max_iso = today

    jurisdiction = project.jurisdiction or "California"
    rows = []
    for analyte, matrix in sorted(pairs):
        thresholds = (
            Threshold.query
            .filter(
                Threshold.analyte == analyte,
                Threshold.matrix == matrix,
                Threshold.jurisdiction.in_([jurisdiction, "Both"]),
                or_(Threshold.effective_date.is_(None),
                    Threshold.effective_date <= max_iso),
                or_(Threshold.superseded_date.is_(None),
                    Threshold.superseded_date > min_iso),
            )
            .order_by(Threshold.threshold_name.asc())
            .all()
        )
        for t in thresholds:
            rows.append([
                t.analyte, t.matrix, t.threshold_name, t.regulatory_body,
                t.value, t.units, t.threshold_type, t.jurisdiction,
                t.effective_date, t.superseded_date, t.source_citation,
            ])
    return rows
